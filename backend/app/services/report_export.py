"""
PDF and Excel exporters for user lifestyle reports.
"""

from __future__ import annotations

import io
import re
from decimal import ROUND_HALF_UP, Decimal, InvalidOperation
from typing import Any, Dict, List, NamedTuple, Optional, Tuple

from fpdf import FPDF
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from app.services.habit_score import format_habit_score

# Excel number formats that keep cells numeric while preserving the canonical
# display used elsewhere in the app.
_XL_HABIT_SCORE_FORMAT = '0"/100"'
_XL_WEIGHT_FORMAT = "0%"
_XL_DECIMAL_FORMAT = "0.0"
_XL_INTEGER_FORMAT = "0"


def _is_habit_score_key(key: str) -> bool:
    return key == "habit_score" or key == "current_habit_score" or key.endswith(
        "_habit_score"
    )


def _round_half_up(value: float, digits: int) -> float:
    """Round using decimal half-up so 0.35 -> 0.4 and 92.85 -> 92.9."""
    try:
        quantum = Decimal(1).scaleb(-digits)
        return float(Decimal(str(value)).quantize(quantum, rounding=ROUND_HALF_UP))
    except (InvalidOperation, ValueError):
        return round(value, digits)


def _fmt_weight(value: Any) -> str:
    """Habit-score weights are fractions of 1.0 — render them as percentages."""
    if isinstance(value, (int, float)) and not isinstance(value, bool):
        return f"{_round_half_up(float(value) * 100, 0):.0f}%"
    return _fmt(value)


def _fmt(value: Any, *, key: Optional[str] = None) -> str:
    if key and _is_habit_score_key(key):
        return format_habit_score(value, empty="-")
    if value is None:
        return "-"
    if isinstance(value, bool):
        return "Yes" if value else "No"
    if isinstance(value, float):
        if abs(value) >= 0.1 or value == 0:
            return f"{_round_half_up(value, 1):.1f}"
        return f"{_round_half_up(value, 2):.2f}"
    if isinstance(value, dict):
        return ", ".join(f"{k}: {_fmt(v, key=k)}" for k, v in value.items())
    if isinstance(value, list):
        if not value:
            return "-"
        if all(not isinstance(x, (dict, list)) for x in value):
            return ", ".join(_fmt(x) for x in value)
        return f"{len(value)} items"
    return str(value)


class SummaryRow(NamedTuple):
    """A summary metric with a text form (PDF) and a numeric form (Excel)."""

    label: str
    text: str
    number: Optional[float]
    number_format: Optional[str]


def _numeric_cell(value: Any, key: Optional[str]) -> Tuple[Optional[float], Optional[str]]:
    if isinstance(value, bool) or not isinstance(value, (int, float)):
        return None, None
    if key and _is_habit_score_key(key):
        return float(value), _XL_HABIT_SCORE_FORMAT
    if isinstance(value, int):
        return value, _XL_INTEGER_FORMAT
    return _round_half_up(float(value), 2), _XL_DECIMAL_FORMAT


def _summary_rows(summary: Dict[str, Any]) -> List[SummaryRow]:
    rows: List[SummaryRow] = []
    for key, value in summary.items():
        label = key.replace("_", " ").title()
        if isinstance(value, dict) and key in ("breakdown", "weights", "trend"):
            for sub_key, sub_value in value.items():
                sub_label = f"{label} / {sub_key.replace('_', ' ').title()}"
                if key == "weights":
                    number = (
                        float(sub_value)
                        if isinstance(sub_value, (int, float))
                        and not isinstance(sub_value, bool)
                        else None
                    )
                    rows.append(
                        SummaryRow(
                            sub_label,
                            _fmt_weight(sub_value),
                            number,
                            _XL_WEIGHT_FORMAT if number is not None else None,
                        )
                    )
                else:
                    number, fmt = _numeric_cell(sub_value, sub_key)
                    rows.append(
                        SummaryRow(sub_label, _fmt(sub_value, key=sub_key), number, fmt)
                    )
        else:
            number, fmt = _numeric_cell(value, key)
            rows.append(SummaryRow(label, _fmt(value, key=key), number, fmt))
    return rows


def _flatten_summary(summary: Dict[str, Any]) -> List[Tuple[str, str]]:
    return [(row.label, row.text) for row in _summary_rows(summary)]


# fpdf's built-in fonts are latin-1 only, but insight and recommendation copy
# is written with typographic punctuation. Transliterate the characters the
# services actually emit so the text stays readable.
_PDF_TRANSLITERATIONS = {
    "\u2014": "-",    # em dash
    "\u2013": "-",    # en dash
    "\u2018": "'",
    "\u2019": "'",
    "\u201c": '"',
    "\u201d": '"',
    "\u2026": "...",
    "\u2192": "->",
    "\u2190": "<-",
    "\u2264": "<=",
    "\u2265": ">=",
    "\u2248": "~",
    "\u2022": "-",
    "\u00a0": " ",
}


def _core_font_safe(text: str) -> str:
    """Make ``text`` renderable by the latin-1 core fonts.

    Without this an em dash anywhere in the payload aborts the whole export
    with ``FPDFUnicodeEncodingException``; anything unmapped degrades to "?"
    rather than failing the download.
    """
    for source, target in _PDF_TRANSLITERATIONS.items():
        text = text.replace(source, target)
    return text.encode("latin-1", "replace").decode("latin-1")


class ReportPDF(FPDF):
    def normalize_text(self, txt):
        # Single choke point: every cell/multi_cell/write goes through here.
        return super().normalize_text(_core_font_safe(str(txt)))

    def header(self):
        self.set_font("Helvetica", "B", 10)
        self.set_text_color(80, 80, 80)
        self.cell(0, 6, "ICAP - Intelligent Cognitive Alarm", align="L")
        self.ln(8)

    def footer(self):
        self.set_y(-12)
        self.set_font("Helvetica", "I", 8)
        self.set_text_color(120, 120, 120)
        self.cell(0, 8, f"Page {self.page_no()}", align="C")


def render_pdf(report: Dict[str, Any]) -> bytes:
    """Render a report payload to PDF bytes."""
    pdf = ReportPDF()
    pdf.set_auto_page_break(auto=True, margin=15)
    pdf.add_page()

    pdf.set_font("Helvetica", "B", 18)
    pdf.set_text_color(15, 23, 42)
    pdf.cell(0, 10, report.get("title", "Report"), new_x="LMARGIN", new_y="NEXT")

    pdf.set_font("Helvetica", "", 10)
    pdf.set_text_color(71, 85, 105)
    period = report.get("period") or {}
    pdf.multi_cell(
        0,
        5,
        (
            f"Period: {period.get('start_date', '-')} to {period.get('end_date', '-')} "
            f"({period.get('days', '-')} days)\n"
            f"Generated: {report.get('generated_at', '-')}"
        ),
    )
    pdf.ln(2)

    if report.get("description"):
        pdf.set_font("Helvetica", "I", 10)
        pdf.multi_cell(0, 5, report["description"])
        pdf.ln(2)

    if report.get("is_empty"):
        pdf.set_fill_color(254, 243, 199)
        pdf.set_font("Helvetica", "B", 11)
        pdf.set_text_color(146, 64, 14)
        pdf.multi_cell(
            0,
            8,
            report.get("empty_message")
            or "No data available for this period. Complete a verified wake-up to unlock this report.",
            fill=True,
        )
        pdf.ln(4)

    sections = report.get("sections") or {}
    summary = sections.get("summary") or {}
    if summary:
        pdf.set_font("Helvetica", "B", 13)
        pdf.set_text_color(15, 23, 42)
        pdf.cell(0, 8, "Summary", new_x="LMARGIN", new_y="NEXT")
        pdf.set_draw_color(37, 99, 235)
        pdf.set_line_width(0.4)
        y = pdf.get_y()
        pdf.line(pdf.l_margin, y, pdf.w - pdf.r_margin, y)
        pdf.ln(3)

        pdf.set_font("Helvetica", "", 10)
        pdf.set_text_color(30, 41, 59)
        for label, value in _flatten_summary(summary):
            # Write full-width lines to avoid cell/multi_cell width conflicts
            pdf.multi_cell(0, 6, f"{label}: {value}")
            pdf.ln(1)

    insights = report.get("insights") or []
    if insights:
        pdf.ln(4)
        pdf.set_font("Helvetica", "B", 13)
        pdf.set_text_color(15, 23, 42)
        pdf.cell(0, 8, "Insights", new_x="LMARGIN", new_y="NEXT")
        pdf.set_font("Helvetica", "", 10)
        pdf.set_text_color(30, 41, 59)
        for item in insights:
            text = item if isinstance(item, str) else _fmt(item)
            pdf.multi_cell(0, 5, f"- {text}")
            pdf.ln(1)

    # Extra detail tables for common breakdowns
    _pdf_detail_tables(pdf, sections)
    _pdf_generic_tables(pdf, sections.get("tables") or [])

    out = pdf.output()
    if isinstance(out, (bytes, bytearray)):
        return bytes(out)
    return out.encode("latin-1")


def _pdf_detail_tables(pdf: ReportPDF, sections: Dict[str, Any]) -> None:
    challenge = sections.get("challenge_performance") or {}
    by_type = challenge.get("by_type") or {}
    if by_type:
        pdf.ln(3)
        pdf.set_font("Helvetica", "B", 12)
        pdf.cell(0, 7, "Challenge Types", new_x="LMARGIN", new_y="NEXT")
        pdf.set_font("Helvetica", "B", 9)
        for header, width in (
            ("Type", 40),
            ("Attempts", 25),
            ("Accuracy %", 30),
            ("Avg Time", 25),
            ("Points", 25),
        ):
            pdf.cell(width, 6, header, border=1)
        pdf.ln()
        pdf.set_font("Helvetica", "", 9)
        for ctype, stats in by_type.items():
            pdf.cell(40, 6, str(ctype)[:18], border=1)
            pdf.cell(25, 6, str(stats.get("total", 0)), border=1)
            pdf.cell(30, 6, _fmt(stats.get("accuracy")), border=1)
            pdf.cell(25, 6, _fmt(stats.get("avg_time")), border=1)
            pdf.cell(25, 6, str(stats.get("points", 0)), border=1)
            pdf.ln()

    wake_stats = sections.get("wake_stats") or {}
    by_weekday = wake_stats.get("by_weekday") or []
    if by_weekday:
        pdf.ln(3)
        pdf.set_font("Helvetica", "B", 12)
        pdf.cell(0, 7, "Wakes by Weekday", new_x="LMARGIN", new_y="NEXT")
        pdf.set_font("Helvetica", "", 10)
        for row in by_weekday:
            pdf.cell(0, 5, f"{row.get('weekday', '-')}: {row.get('count', 0)}", new_x="LMARGIN", new_y="NEXT")

    # Distribution maps (admin system reports)
    for key, title in (
        ("role_distribution", "Role Distribution"),
        ("type_distribution", "Alarm Type Distribution"),
        ("challenge_type_distribution", "Challenge Type Distribution"),
        ("difficulty_distribution", "Difficulty Distribution"),
    ):
        dist = sections.get(key)
        if isinstance(dist, dict) and dist and all(
            not isinstance(v, dict) for v in dist.values()
        ):
            pdf.ln(3)
            pdf.set_font("Helvetica", "B", 12)
            pdf.cell(0, 7, title, new_x="LMARGIN", new_y="NEXT")
            pdf.set_font("Helvetica", "", 10)
            for label, value in dist.items():
                pdf.cell(
                    0,
                    5,
                    f"{str(label).replace('_', ' ').title()}: {_fmt(value)}",
                    new_x="LMARGIN",
                    new_y="NEXT",
                )


def _pdf_generic_tables(pdf: ReportPDF, tables: List[Any]) -> None:
    """Render generic ``sections.tables`` entries used by admin system reports."""
    for table in tables:
        if not isinstance(table, dict):
            continue
        headers = table.get("headers") or []
        rows = table.get("rows") or []
        if not headers or not rows:
            continue
        title = table.get("title") or "Details"
        pdf.ln(3)
        pdf.set_font("Helvetica", "B", 12)
        pdf.cell(0, 7, str(title)[:60], new_x="LMARGIN", new_y="NEXT")

        usable = pdf.w - pdf.l_margin - pdf.r_margin
        col_count = max(len(headers), 1)
        col_w = max(18.0, usable / col_count)
        pdf.set_font("Helvetica", "B", 8)
        for header in headers:
            pdf.cell(col_w, 6, str(header)[:22], border=1)
        pdf.ln()
        pdf.set_font("Helvetica", "", 8)
        for row in rows[:40]:
            cells = list(row) if isinstance(row, (list, tuple)) else [row]
            for i in range(col_count):
                val = cells[i] if i < len(cells) else ""
                pdf.cell(col_w, 6, _fmt(val)[:22], border=1)
            pdf.ln()


def render_excel(report: Dict[str, Any]) -> bytes:
    """Render a report payload to Excel (.xlsx) bytes."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Summary"

    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill("solid", fgColor="2563EB")
    title_font = Font(bold=True, size=14)
    warn_fill = PatternFill("solid", fgColor="FEF3C7")

    ws["A1"] = report.get("title", "Report")
    ws["A1"].font = title_font
    ws.merge_cells("A1:B1")

    period = report.get("period") or {}
    ws["A2"] = "Period start"
    ws["B2"] = period.get("start_date")
    ws["A3"] = "Period end"
    ws["B3"] = period.get("end_date")
    ws["A4"] = "Days"
    ws["B4"] = period.get("days")
    ws["A5"] = "Generated at"
    ws["B5"] = report.get("generated_at")
    ws["A6"] = "Empty report"
    ws["B6"] = bool(report.get("is_empty"))

    row = 8
    if report.get("is_empty"):
        ws.cell(
            row=row,
            column=1,
            value=report.get("empty_message")
            or "No data available for this period. Complete a verified wake-up to unlock this report.",
        )
        ws.cell(row=row, column=1).fill = warn_fill
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=2)
        row += 2

    ws.cell(row=row, column=1, value="Metric").font = header_font
    ws.cell(row=row, column=1).fill = header_fill
    ws.cell(row=row, column=2, value="Value").font = header_font
    ws.cell(row=row, column=2).fill = header_fill
    row += 1

    sections = report.get("sections") or {}
    summary = sections.get("summary") or {}
    for item in _summary_rows(summary):
        ws.cell(row=row, column=1, value=item.label)
        cell = ws.cell(row=row, column=2)
        if item.number is None:
            cell.value = item.text
        else:
            cell.value = item.number
            cell.number_format = item.number_format or _XL_DECIMAL_FORMAT
        row += 1

    # Insights sheet
    insights = report.get("insights") or []
    ws2 = wb.create_sheet("Insights")
    ws2["A1"] = "Insights"
    ws2["A1"].font = title_font
    if not insights:
        ws2["A2"] = "No insights for this period."
    else:
        for i, item in enumerate(insights, start=2):
            ws2.cell(row=i, column=1, value=item if isinstance(item, str) else _fmt(item))

    # Detail sheets
    challenge = sections.get("challenge_performance") or {}
    by_type = challenge.get("by_type") or {}
    if by_type:
        ws3 = wb.create_sheet("Challenge Types")
        headers = ["Type", "Total", "Correct", "Accuracy", "Avg Time", "Points"]
        for col, h in enumerate(headers, start=1):
            cell = ws3.cell(row=1, column=col, value=h)
            cell.font = header_font
            cell.fill = header_fill
        for i, (ctype, stats) in enumerate(by_type.items(), start=2):
            ws3.cell(row=i, column=1, value=ctype)
            ws3.cell(row=i, column=2, value=stats.get("total", 0))
            ws3.cell(row=i, column=3, value=stats.get("correct", 0))
            ws3.cell(row=i, column=4, value=stats.get("accuracy", 0))
            ws3.cell(row=i, column=5, value=stats.get("avg_time", 0))
            ws3.cell(row=i, column=6, value=stats.get("points", 0))

    wake_stats = sections.get("wake_stats") or {}
    by_weekday = wake_stats.get("by_weekday") or []
    if by_weekday:
        ws4 = wb.create_sheet("Wake Weekdays")
        ws4.cell(row=1, column=1, value="Weekday").font = header_font
        ws4.cell(row=1, column=1).fill = header_fill
        ws4.cell(row=1, column=2, value="Count").font = header_font
        ws4.cell(row=1, column=2).fill = header_fill
        for i, item in enumerate(by_weekday, start=2):
            ws4.cell(row=i, column=1, value=item.get("weekday"))
            ws4.cell(row=i, column=2, value=item.get("count", 0))

    habit_trends = (sections.get("trends") or sections.get("habit_trends") or {})
    series = habit_trends.get("series") or []
    if series:
        ws5 = wb.create_sheet("Daily Series")
        keys = list(series[0].keys())
        for col, key in enumerate(keys, start=1):
            cell = ws5.cell(row=1, column=col, value=key)
            cell.font = header_font
            cell.fill = header_fill
        for r_i, item in enumerate(series, start=2):
            for c_i, key in enumerate(keys, start=1):
                val = item.get(key)
                if isinstance(val, (dict, list)):
                    val = _fmt(val)
                ws5.cell(row=r_i, column=c_i, value=val)

    # Generic admin tables
    for table in sections.get("tables") or []:
        if not isinstance(table, dict):
            continue
        headers = table.get("headers") or []
        rows = table.get("rows") or []
        if not headers:
            continue
        title = str(table.get("title") or "Details")[:31]
        # Avoid duplicate sheet names
        base = title
        suffix = 1
        while title in wb.sheetnames:
            title = f"{base[:28]}_{suffix}"
            suffix += 1
        ws_t = wb.create_sheet(title)
        for col, h in enumerate(headers, start=1):
            cell = ws_t.cell(row=1, column=col, value=h)
            cell.font = header_font
            cell.fill = header_fill
        for r_i, row in enumerate(rows, start=2):
            cells = list(row) if isinstance(row, (list, tuple)) else [row]
            for c_i, val in enumerate(cells, start=1):
                if isinstance(val, (dict, list)):
                    val = _fmt(val)
                ws_t.cell(row=r_i, column=c_i, value=val)

    for sheet in wb.worksheets:
        for col in range(1, min(sheet.max_column, 12) + 1):
            sheet.column_dimensions[get_column_letter(col)].width = 22
            for cell in sheet[get_column_letter(col)]:
                cell.alignment = Alignment(vertical="center", wrap_text=True)

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


#: Everything outside this set is replaced in a download filename. Keeps path
#: separators, quotes and CR/LF out of the ``Content-Disposition`` header, so a
#: report field can never break the header or suggest a path to the browser.
_UNSAFE_FILENAME_CHARS = re.compile(r"[^A-Za-z0-9._-]+")


def safe_attachment_filename(name: str, *, fallback: str = "icap_report") -> str:
    """Reduce ``name`` to a filename that is safe to put in a header."""
    cleaned = _UNSAFE_FILENAME_CHARS.sub("_", str(name or "")).strip("._")
    # A leading dot or a bare extension is not a usable download name.
    return cleaned[:120] if cleaned and not cleaned.startswith(".") else fallback


def content_disposition(filename: str) -> str:
    """Build an ``attachment`` header value that cannot be injected into."""
    return f'attachment; filename="{safe_attachment_filename(filename)}"'


def export_report(report: Dict[str, Any], fmt: str) -> Tuple[bytes, str, str]:
    """Return ``(content, media_type, filename)`` for pdf or excel."""
    report_type = report.get("report_type", "report")
    period = report.get("period") or {}
    start = period.get("start_date", "start")
    end = period.get("end_date", "end")
    fmt = (fmt or "pdf").lower()
    if fmt in ("xlsx", "excel", "xls"):
        content = render_excel(report)
        filename = safe_attachment_filename(
            f"icap_{report_type}_report_{start}_to_{end}.xlsx"
        )
        return content, (
            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        ), filename
    if fmt != "pdf":
        raise ValueError("format must be 'pdf' or 'excel'")
    content = render_pdf(report)
    filename = safe_attachment_filename(
        f"icap_{report_type}_report_{start}_to_{end}.pdf"
    )
    return content, "application/pdf", filename
