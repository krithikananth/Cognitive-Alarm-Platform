"""Tests for admin System Reports (User / Alarm / Habit / Platform + export)."""

from datetime import datetime, time, timedelta, timezone
from io import BytesIO

from openpyxl import load_workbook

from app.models.alarm import Alarm, AlarmChallengeLog, AlarmType, ChallengeType
from app.models.alarm_wake_event import AlarmWakeEvent
from app.models.profile import UserProfile
from app.services.report_export import export_report
from app.services.system_report_service import (
    SystemReportService,
    SystemReportType,
    list_system_report_types,
)


def _make_alarm(db_session, user_id: int) -> Alarm:
    alarm = Alarm(
        user_id=user_id,
        title="System Report Alarm",
        alarm_time=time(7, 0),
        alarm_type=AlarmType.DAILY,
        challenge_type=ChallengeType.MATH,
        challenge_count=1,
        challenge_difficulty="medium",
        snooze_limit=3,
    )
    db_session.add(alarm)
    db_session.commit()
    db_session.refresh(alarm)
    return alarm


def _ensure_profile(db_session, user_id: int, **kwargs) -> UserProfile:
    profile = (
        db_session.query(UserProfile)
        .filter(UserProfile.user_id == user_id)
        .first()
    )
    if profile is None:
        profile = UserProfile(
            user_id=user_id,
            preferred_wake_time=time(7, 0),
            sleep_duration_hours=8.0,
            **kwargs,
        )
        db_session.add(profile)
    else:
        for k, v in kwargs.items():
            setattr(profile, k, v)
    db_session.commit()
    db_session.refresh(profile)
    return profile


def _add_wake(db_session, user_id, alarm_id, dismissed_at, *, verified=True):
    triggered = dismissed_at - timedelta(minutes=5)
    row = AlarmWakeEvent(
        user_id=user_id,
        alarm_id=alarm_id,
        triggered_at=triggered,
        dismissed_at=dismissed_at,
        dismiss_method="challenge",
        snooze_count_at_dismiss=0,
        time_to_dismiss_seconds=300,
        verified=verified,
        wakefulness_score=78.0,
        wakefulness_level="alert",
        failed_attempts=0,
        challenges_required=1,
        challenges_completed=1,
    )
    db_session.add(row)
    db_session.commit()
    return row


def _add_challenge(db_session, user_id, alarm_id, created_at, *, correct=True):
    row = AlarmChallengeLog(
        user_id=user_id,
        alarm_id=alarm_id,
        challenge_type="math",
        difficulty="medium",
        challenge_prompt="2+2?",
        is_correct=correct,
        time_taken_seconds=12,
        points_earned=10 if correct else 0,
        created_at=created_at,
    )
    db_session.add(row)
    db_session.commit()
    return row


class TestSystemReportService:
    def test_list_types(self):
        types = list_system_report_types()
        assert {t["type"] for t in types} == {"user", "alarm", "habit", "platform"}

    def test_user_report_emptyish(self, db_session, test_user):
        payload = SystemReportService.build_report(
            db_session, SystemReportType.USER, days=30
        )
        assert payload["report_type"] == "user"
        assert "summary" in payload["sections"]
        assert payload["sections"]["summary"]["total_users"] >= 1

    def test_all_types_with_activity(self, db_session, test_user):
        _ensure_profile(db_session, test_user.id, streak_days=3, best_streak=5)
        alarm = _make_alarm(db_session, test_user.id)
        now = datetime.now(timezone.utc)
        _add_wake(db_session, test_user.id, alarm.id, now - timedelta(days=1))
        _add_challenge(
            db_session, test_user.id, alarm.id, now - timedelta(days=1)
        )

        for rt in SystemReportType:
            payload = SystemReportService.build_report(db_session, rt, days=30)
            assert payload["report_type"] == rt.value
            assert payload["period"]["days"] == 30
            assert "summary" in payload["sections"]
            assert isinstance(payload["insights"], list)

        user_report = SystemReportService.build_report(
            db_session, SystemReportType.USER, days=30
        )
        assert user_report["sections"]["summary"]["engaged_users"] >= 1

        alarm_report = SystemReportService.build_report(
            db_session, SystemReportType.ALARM, days=30
        )
        assert alarm_report["sections"]["summary"]["wake_events"] >= 1

        habit_report = SystemReportService.build_report(
            db_session, SystemReportType.HABIT, days=30
        )
        assert habit_report["sections"]["summary"]["total_profiles"] >= 1

        platform_report = SystemReportService.build_report(
            db_session, SystemReportType.PLATFORM, days=30
        )
        assert platform_report["sections"]["summary"]["wake_events"] >= 1

    def test_date_range(self, db_session, test_user):
        from datetime import date

        end = date.today()
        start = end - timedelta(days=6)
        payload = SystemReportService.build_report(
            db_session,
            SystemReportType.PLATFORM,
            start_date=start,
            end_date=end,
        )
        assert payload["period"]["days"] == 7
        assert payload["period"]["start_date"] == start.isoformat()

    def test_export_pdf_and_excel(self, db_session, test_user):
        _ensure_profile(db_session, test_user.id)
        alarm = _make_alarm(db_session, test_user.id)
        now = datetime.now(timezone.utc)
        _add_wake(db_session, test_user.id, alarm.id, now - timedelta(hours=2))

        payload = SystemReportService.build_report(
            db_session, SystemReportType.USER, days=14
        )
        pdf_bytes, pdf_mime, pdf_name = export_report(payload, "pdf")
        assert pdf_mime == "application/pdf"
        assert pdf_name.endswith(".pdf")
        assert pdf_bytes[:4] == b"%PDF"

        xlsx_bytes, xlsx_mime, xlsx_name = export_report(payload, "excel")
        assert "spreadsheet" in xlsx_mime
        assert xlsx_name.endswith(".xlsx")
        wb = load_workbook(BytesIO(xlsx_bytes))
        assert "Summary" in wb.sheetnames


class TestSystemReportsAPI:
    def test_list_requires_admin(self, client, auth_headers):
        res = client.get("/api/v1/admin/system-reports", headers=auth_headers)
        assert res.status_code == 403

    def test_list_and_get_as_admin(self, client, admin_headers, db_session, test_user):
        _ensure_profile(db_session, test_user.id)
        list_res = client.get("/api/v1/admin/system-reports", headers=admin_headers)
        assert list_res.status_code == 200
        types = list_res.json()["reports"]
        assert len(types) == 4

        for rt in ("user", "alarm", "habit", "platform"):
            res = client.get(
                f"/api/v1/admin/system-reports/{rt}",
                headers=admin_headers,
                params={"days": 30},
            )
            assert res.status_code == 200, res.text
            body = res.json()
            assert body["report_type"] == rt
            assert "sections" in body

    def test_export_pdf(self, client, admin_headers):
        res = client.get(
            "/api/v1/admin/system-reports/platform/export",
            headers=admin_headers,
            params={"format": "pdf", "days": 7},
        )
        assert res.status_code == 200
        assert res.headers["content-type"].startswith("application/pdf")
        assert res.content[:4] == b"%PDF"

    def test_export_excel(self, client, admin_headers):
        res = client.get(
            "/api/v1/admin/system-reports/alarm/export",
            headers=admin_headers,
            params={"format": "excel", "days": 7},
        )
        assert res.status_code == 200
        assert "spreadsheet" in res.headers["content-type"]
        wb = load_workbook(BytesIO(res.content))
        assert "Summary" in wb.sheetnames

    def test_unknown_type(self, client, admin_headers):
        res = client.get(
            "/api/v1/admin/system-reports/unknown",
            headers=admin_headers,
        )
        assert res.status_code == 404

    def test_health_reports_still_works(self, client, admin_headers):
        res = client.get("/api/v1/admin/reports", headers=admin_headers)
        assert res.status_code == 200
        assert "system" in res.json()
