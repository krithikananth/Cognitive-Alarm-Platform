"""Tests for lifestyle Reports (JSON + PDF/Excel export)."""

from datetime import date, datetime, time, timedelta, timezone
from io import BytesIO

import pytest
from openpyxl import load_workbook

from app.models.alarm import Alarm, AlarmChallengeLog, AlarmType, ChallengeType
from app.models.alarm_wake_event import AlarmWakeEvent
from app.models.profile import UserProfile
from app.services.report_export import export_report, render_excel, render_pdf
from app.services.report_service import ReportService, ReportType, resolve_date_window


def _make_alarm(db_session, user_id: int) -> Alarm:
    alarm = Alarm(
        user_id=user_id,
        title="Report Alarm",
        alarm_time=time(7, 0),
        alarm_type=AlarmType.DAILY,
        challenge_type=ChallengeType.MATH,
        challenge_count=1,
        challenge_difficulty="medium",
        snooze_limit=3,
    )
    db_session.add(alarm)
    db_session.commit()
    db_session.refresh(alarm)
    return alarm


def _ensure_profile(db_session, user_id: int, **kwargs) -> UserProfile:
    profile = (
        db_session.query(UserProfile)
        .filter(UserProfile.user_id == user_id)
        .first()
    )
    if profile is None:
        profile = UserProfile(
            user_id=user_id,
            preferred_wake_time=time(7, 0),
            sleep_duration_hours=8.0,
            **kwargs,
        )
        db_session.add(profile)
    else:
        for k, v in kwargs.items():
            setattr(profile, k, v)
    db_session.commit()
    db_session.refresh(profile)
    return profile


def _add_wake(db_session, user_id, alarm_id, dismissed_at, *, verified=True, snoozes=0):
    triggered = dismissed_at - timedelta(minutes=5)
    row = AlarmWakeEvent(
        user_id=user_id,
        alarm_id=alarm_id,
        triggered_at=triggered,
        dismissed_at=dismissed_at,
        dismiss_method="challenge",
        snooze_count_at_dismiss=snoozes,
        time_to_dismiss_seconds=300,
        verified=verified,
        wakefulness_score=78.0,
        wakefulness_level="alert",
        failed_attempts=0,
        challenges_required=1,
        challenges_completed=1,
    )
    db_session.add(row)
    db_session.commit()
    return row


def _add_challenge(db_session, user_id, alarm_id, created_at, *, correct=True):
    row = AlarmChallengeLog(
        user_id=user_id,
        alarm_id=alarm_id,
        challenge_type="math",
        difficulty="medium",
        challenge_prompt="2+2?",
        is_correct=correct,
        time_taken_seconds=12,
        points_earned=10 if correct else 0,
        created_at=created_at,
    )
    db_session.add(row)
    db_session.commit()
    return row


class TestResolveDateWindow:
    def test_days_lookback(self):
        start, end, days = resolve_date_window(days=14)
        assert days == 14
        assert end >= start

    def test_explicit_range(self):
        start, end, days = resolve_date_window(
            start_date=date(2026, 7, 1),
            end_date=date(2026, 7, 10),
        )
        assert days == 10
        assert start.date() == date(2026, 7, 1)
        assert end.date() == date(2026, 7, 10)

    def test_invalid_range(self):
        with pytest.raises(ValueError):
            resolve_date_window(start_date=date(2026, 7, 10), end_date=date(2026, 7, 1))

    def test_partial_range_rejected(self):
        with pytest.raises(ValueError):
            resolve_date_window(start_date=date(2026, 7, 1))


class TestReportService:
    def test_empty_reports(self, db_session, test_user):
        _ensure_profile(db_session, test_user.id)
        for rt in ReportType:
            report = ReportService.build_report(
                db_session, test_user.id, rt, days=30
            )
            assert report["is_empty"] is True
            assert report["empty_message"]
            assert report["report_type"] == rt.value
            assert "summary" in report["sections"]

    def test_populated_reports_reuse_metrics(self, db_session, test_user):
        _ensure_profile(db_session, test_user.id)
        alarm = _make_alarm(db_session, test_user.id)
        now = datetime.now(timezone.utc)
        for i in range(3):
            ts = now - timedelta(days=i)
            wake_at = ts.replace(hour=7, minute=5, second=0, microsecond=0)
            _add_wake(db_session, test_user.id, alarm.id, wake_at)
            _add_challenge(
                db_session,
                test_user.id,
                alarm.id,
                wake_at + timedelta(minutes=1),
                correct=True,
            )

        wake_report = ReportService.build_report(
            db_session, test_user.id, ReportType.WAKE, days=30
        )
        assert wake_report["is_empty"] is False
        assert wake_report["sections"]["summary"]["verified_wakes"] == 3

        challenge_report = ReportService.build_report(
            db_session, test_user.id, ReportType.CHALLENGE, days=30
        )
        assert challenge_report["is_empty"] is False
        assert challenge_report["sections"]["summary"]["total_attempts"] == 3
        assert challenge_report["sections"]["summary"]["accuracy"] == 100.0

        habit_report = ReportService.build_report(
            db_session, test_user.id, ReportType.HABIT, days=30
        )
        assert habit_report["is_empty"] is False
        assert "habit_score" in habit_report["sections"]["summary"]

        prod = ReportService.build_report(
            db_session, test_user.id, ReportType.PRODUCTIVITY, days=30
        )
        assert prod["is_empty"] is False
        assert prod["sections"]["summary"]["verified_wakes"] == 3

        sleep = ReportService.build_report(
            db_session, test_user.id, ReportType.SLEEP, days=30
        )
        assert sleep["is_empty"] is False
        assert sleep["sections"]["summary"]["observed_days"] >= 1

    def test_date_filter_excludes_old_events(self, db_session, test_user):
        _ensure_profile(db_session, test_user.id)
        alarm = _make_alarm(db_session, test_user.id)
        now = datetime.now(timezone.utc)
        old = now - timedelta(days=40)
        _add_wake(
            db_session,
            test_user.id,
            alarm.id,
            old.replace(hour=7, minute=0, second=0, microsecond=0),
        )
        recent = now - timedelta(days=2)
        _add_wake(
            db_session,
            test_user.id,
            alarm.id,
            recent.replace(hour=7, minute=0, second=0, microsecond=0),
        )

        report = ReportService.build_report(
            db_session, test_user.id, ReportType.WAKE, days=7
        )
        assert report["sections"]["summary"]["total_wake_events"] == 1

        ranged = ReportService.build_report(
            db_session,
            test_user.id,
            ReportType.WAKE,
            start_date=(now - timedelta(days=45)).date(),
            end_date=now.date(),
        )
        assert ranged["sections"]["summary"]["total_wake_events"] == 2


class TestReportExport:
    def test_pdf_and_excel_empty(self, db_session, test_user):
        _ensure_profile(db_session, test_user.id)
        report = ReportService.build_report(
            db_session, test_user.id, ReportType.HABIT, days=30
        )
        pdf_bytes = render_pdf(report)
        assert pdf_bytes[:4] == b"%PDF"
        assert b"No data available" in pdf_bytes or b"Empty" in pdf_bytes or len(pdf_bytes) > 200

        xlsx_bytes = render_excel(report)
        assert xlsx_bytes[:2] == b"PK"
        wb = load_workbook(BytesIO(xlsx_bytes))
        assert "Summary" in wb.sheetnames
        assert wb["Summary"]["B6"].value is True  # is_empty

    def test_pdf_and_excel_with_data(self, db_session, test_user):
        _ensure_profile(db_session, test_user.id)
        alarm = _make_alarm(db_session, test_user.id)
        now = datetime.now(timezone.utc)
        _add_wake(
            db_session,
            test_user.id,
            alarm.id,
            now.replace(hour=7, minute=2, second=0, microsecond=0),
        )
        _add_challenge(
            db_session,
            test_user.id,
            alarm.id,
            now.replace(hour=7, minute=3, second=0, microsecond=0),
        )

        report = ReportService.build_report(
            db_session, test_user.id, ReportType.CHALLENGE, days=30
        )
        content, media, filename = export_report(report, "pdf")
        assert media == "application/pdf"
        assert filename.endswith(".pdf")
        assert content[:4] == b"%PDF"

        content, media, filename = export_report(report, "excel")
        assert "spreadsheetml" in media
        assert filename.endswith(".xlsx")
        wb = load_workbook(BytesIO(content))
        assert "Summary" in wb.sheetnames
        assert "Challenge Types" in wb.sheetnames

    def test_pdf_export_survives_typographic_punctuation(self, db_session, test_user):
        """Insight copy contains em dashes; the core PDF fonts are latin-1 only.

        ``BehavioralAnalyticsService`` emits lines such as "trending down — keep
        reinforcing...", which used to abort the whole export with
        ``FPDFUnicodeEncodingException``.
        """
        _ensure_profile(db_session, test_user.id)
        report = ReportService.build_report(
            db_session, test_user.id, ReportType.HABIT, days=30
        )
        report["insights"] = [
            "Snooze volume is trending down \u2014 keep reinforcing dismissals.",
            "Wake time \u2192 07:00 \u00b1 15 min, consistency \u2248 82%.",
        ]

        pdf_bytes = render_pdf(report)

        assert pdf_bytes[:4] == b"%PDF"
        assert len(pdf_bytes) > 200


class TestReportsAPI:
    def test_list_report_types(self, client, auth_headers):
        res = client.get("/api/v1/reports", headers=auth_headers)
        assert res.status_code == 200
        types = {r["type"] for r in res.json()["reports"]}
        assert types == {"habit", "wake", "challenge", "productivity", "sleep"}

    def test_get_each_report_type(self, client, auth_headers):
        for rt in ("habit", "wake", "challenge", "productivity", "sleep"):
            res = client.get(f"/api/v1/reports/{rt}?days=30", headers=auth_headers)
            assert res.status_code == 200, rt
            body = res.json()
            assert body["report_type"] == rt
            assert body["is_empty"] is True
            assert "period" in body

    def test_date_filter_validation(self, client, auth_headers):
        res = client.get(
            "/api/v1/reports/wake?start_date=2026-07-10&end_date=2026-07-01",
            headers=auth_headers,
        )
        assert res.status_code == 400

    def test_unknown_type(self, client, auth_headers):
        res = client.get("/api/v1/reports/unknown", headers=auth_headers)
        assert res.status_code == 404

    def test_export_pdf_and_excel(self, client, auth_headers, db_session, test_user):
        _ensure_profile(db_session, test_user.id)
        alarm = _make_alarm(db_session, test_user.id)
        now = datetime.now(timezone.utc)
        _add_wake(
            db_session,
            test_user.id,
            alarm.id,
            now.replace(hour=7, minute=0, second=0, microsecond=0),
        )

        pdf = client.get(
            "/api/v1/reports/wake/export?format=pdf&days=30",
            headers=auth_headers,
        )
        assert pdf.status_code == 200
        assert pdf.headers["content-type"].startswith("application/pdf")
        assert pdf.content[:4] == b"%PDF"
        assert "attachment" in pdf.headers.get("content-disposition", "")

        xlsx = client.get(
            "/api/v1/reports/wake/export?format=excel&days=30",
            headers=auth_headers,
        )
        assert xlsx.status_code == 200
        assert xlsx.content[:2] == b"PK"
        wb = load_workbook(BytesIO(xlsx.content))
        assert "Summary" in wb.sheetnames

    def test_export_empty_still_valid_files(self, client, auth_headers):
        pdf = client.get(
            "/api/v1/reports/sleep/export?format=pdf&days=7",
            headers=auth_headers,
        )
        assert pdf.status_code == 200
        assert pdf.content[:4] == b"%PDF"

        xlsx = client.get(
            "/api/v1/reports/sleep/export?format=xlsx&days=7",
            headers=auth_headers,
        )
        assert xlsx.status_code == 200
        wb = load_workbook(BytesIO(xlsx.content))
        assert wb["Summary"]["B6"].value is True
